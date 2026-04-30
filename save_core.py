import openpyxl.worksheet.worksheet
from PySide6.QtCore import Signal,QThread

from locals import *
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,Border,Side
import os

black_side = Side(border_style="thin", color="000000")
border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)

center_align=Alignment(wrap_text=True,horizontal="center",vertical="center")
title_font=Font(cfg.text_font.value,cfg.text_size.value+6)
font=Font(cfg.text_font.value,cfg.text_size.value)
info_row_height=cfg.text_size.value*3.75
lesson_row_height=cfg.text_size.value*7.5
activity_row_height=cfg.text_size.value*4

grades=list(cfg.grades_info.value.keys())
activities:dict[str,int]={}
last_activities=[]
morning_activity_num=afternoon_activity_num=0
morning_class_num=cfg.morning_class_num.value
afternoon_class_num=cfg.afternoon_class_num.value
lesson_num=morning_class_num+afternoon_class_num
activity_num=len(cfg.activity_info.value)
activity_names=list(cfg.activity_info.value.keys())
activity_names.sort(key=lambda activity: cfg.activity_info.value[activity][0])
for activity in activity_names:
    activity_start_time=cfg.activity_info.value[activity][0]
    f=0
    for lesson,(lesson_start_time,_) in cfg.lessons_time.value.items():
        if activity_start_time<lesson_start_time:
            if int(lesson)<=morning_class_num:
                morning_activity_num+=1
            else:
                afternoon_activity_num+=1
            activities[activity]=int(lesson)
            f=1
            break
    if not f:
        last_activities.append(activity)
last_activities.sort(key=lambda activity: cfg.activity_info.value[activity][0])

teachers:dict[str,set[str]]={}
for grade in grades:
    teachers[grade]=set()
    for class_name in cfg.grades_info.value[grade]:
        clas=results.classes[class_name]
        teachers[grade]|={teacher.name for teacher in clas.teachers.values()}

def column_letter(index):
    """将数字列索引转换为Excel列字母（1->A, 2->B, ..., 27->AA）"""
    result = ""
    while index > 0:
        index -= 1
        result = chr(ord('A') + index % 26) + result
        index //= 26
    return result

def save_grades_timetable(filename:str,ext:str):
    wb=Workbook()
    wb.remove(wb.active)
    for grade in grades:
        ws:openpyxl.worksheet.worksheet.Worksheet=wb.create_sheet(grade+"各班课表")
        ws.column_dimensions["A"].width=cfg.text_size.value-2
        ws.column_dimensions["B"].width=cfg.text_size.value+2
        ws.column_dimensions["C"].width=ws.column_dimensions["D"].width=ws.column_dimensions["E"].width=ws.column_dimensions["F"].width=ws.column_dimensions["G"].width=cfg.text_size.value*1.5
        start_row=1
        for class_name in cfg.grades_info.value[grade]:
            ws.cell(start_row,1).value=cfg.school_name.value
            ws.cell(start_row,1).font=title_font
            ws.cell(start_row,1).border=border
            ws.merge_cells(start_row=start_row,start_column=1,end_row=start_row,end_column=7)
            ws.row_dimensions[start_row].height=ws.row_dimensions[start_row+1].height=ws.row_dimensions[start_row+2].height=info_row_height
            ws.cell(start_row+1,1).value=class_name+"课表"
            ws.cell(start_row+1,1).font=title_font
            ws.cell(start_row+1,1).border=border
            ws.merge_cells(start_row=start_row+1,start_column=1,end_row=start_row+1,end_column=7)
            ws.append(["星期\n节次","","星期一","星期二","星期三","星期四","星期五"])
            for day in range(1,6):
                ws.cell(start_row+2,day+2).border=border
                for row in range(start_row+3,start_row+3+lesson_num+activity_num):
                    ws.row_dimensions[row].height=lesson_row_height

            for lesson in range(1,lesson_num+1):
                ws.cell(start_row+2+lesson,2).border=border
                ws.cell(start_row+2+lesson,2).value=f"{lesson}节\n%02d:%02d~%02d:%02d"%(cfg.lessons_time.value[str(lesson)][0][0],cfg.lessons_time.value[str(lesson)][0][1],cfg.lessons_time.value[str(lesson)][1][0],cfg.lessons_time.value[str(lesson)][1][1])

            timetable=results.classes[class_name].timetable_dataframe
            for day in range(1,6):
                col=day+2
                for lesson in range(1,lesson_num+1):
                    row=start_row+2+lesson
                    time=Time(day,lesson)
                    subject_name=timetable.loc[time.to_str(False,True),time.to_str(True,False)]
                    ws.cell(row,col).value=subject_name
                    ws.cell(row,col).border=border

            activity_row=0
            for activity in list(activities.keys())+last_activities:
                if activity in activities:
                    lesson=activities[activity]
                else:
                    lesson=lesson_num+1
                curr_row=start_row+2+lesson+activity_row
                ws.insert_rows(curr_row)
                ws.cell(curr_row,2).value=activity+" %02d:%02d~%02d:%02d"%(cfg.activity_info.value[activity][0][0],cfg.activity_info.value[activity][0][1],cfg.activity_info.value[activity][1][0],cfg.activity_info.value[activity][1][1])
                ws.cell(curr_row,2).border=border
                ws.row_dimensions[curr_row].height=activity_row_height
                ws.merge_cells(start_row=curr_row,start_column=2,end_row=curr_row,end_column=7)
                activity_row+=1

            ws.merge_cells(start_row=start_row+2,start_column=1,end_row=start_row+2,end_column=2)
            ws.cell(start_row+3,1).value="上午"
            ws.cell(start_row+3,1).border=border
            ws.merge_cells(start_row=start_row+3,start_column=1,end_row=start_row+2+morning_class_num+morning_activity_num,end_column=1)
            afternoon_row=start_row+3+morning_class_num+morning_activity_num
            ws.cell(afternoon_row,1).value="下午"
            ws.cell(afternoon_row,1).border=border
            ws.merge_cells(start_row=afternoon_row,start_column=1,end_row=start_row+2+lesson_num+activity_num,end_column=1)

            start_row+=4+lesson_num+activity_num
        for row in ws[ws.dimensions]:
            for cell in row:
                cell.alignment=center_align
                if cell.font!=title_font:
                    cell.font=font

    wb.save(filename+"-各班课表（年级）"+ext)
    logging.info("保存 各班课表（年级） 成功")

def save_total_grades_timetable(filename:str,ext:str):
    font=Font(cfg.text_font.value,cfg.text_size.value+4)
    total_column=2+(lesson_num+activity_num)*5
    wb=Workbook()
    ws=wb.active
    ws.title=os.path.basename(filename)+"_总表"
    ws["A1"].value=cfg.school_name.value
    ws["A1"].font=title_font
    ws.row_dimensions[1].height=ws.row_dimensions[2].height=info_row_height
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=total_column)

    ws["A2"].value="总课表（班级）"
    ws["A2"].font=title_font
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=total_column)

    ws["A3"].value="班级\n节次\n时间"
    ws.merge_cells("A3:B6")

    for day in range(1,6):
        start_row=7
        start_column=(day-1)*(lesson_num+activity_num)+3
        for grade in grades:
            classes=cfg.grades_info.value[grade]
            ws.cell(start_row,1).value=grade
            ws.merge_cells(start_row=start_row,start_column=1,end_row=start_row+len(classes)-1,end_column=1)
            ws.column_dimensions["B"].width=cfg.text_size.value+3
            for i in range(len(classes)):
                class_name=classes[i]
                ws.cell(start_row+i,2).value=class_name
                clas=results.classes[class_name]
                timetable=clas.timetable_dataframe

                for lesson in range(1,lesson_num+activity_num+1):
                    if lesson<=lesson_num:
                        ws.cell(5,start_column+lesson-1).value=f"{lesson}节"
                        ws.cell(6,start_column+lesson-1).value="%02d:%02d~%02d:%02d"%(cfg.lessons_time.value[str(lesson)][0][0],cfg.lessons_time.value[str(lesson)][0][1],cfg.lessons_time.value[str(lesson)][1][0],cfg.lessons_time.value[str(lesson)][1][1])
                        time=Time(day,lesson)
                        subject_name=timetable.loc[time.to_str(False,True),time.to_str(True,False)]
                        ws.cell(start_row+i,start_column+lesson-1).value=subject_name
                    ws.column_dimensions[column_letter(start_column+lesson-1)].width=cfg.text_size.value+5
            start_row+=len(classes)

        activity_column=0
        for activity in list(activities.keys())+last_activities:
            if activity in activities:
                lesson=activities[activity]
            else:
                lesson=lesson_num+1
            curr_column=start_column+lesson+activity_column-1
            ws.insert_cols(curr_column)
            ws.cell(5,curr_column).value="%02d:%02d~%02d:%02d"%(cfg.activity_info.value[activity][0][0],cfg.activity_info.value[activity][0][1],cfg.activity_info.value[activity][1][0],cfg.activity_info.value[activity][1][1])
            ws.merge_cells(start_row=5,start_column=curr_column,end_row=6,end_column=curr_column)

            ws.cell(7,curr_column).value=activity
            ws.merge_cells(start_row=7,start_column=curr_column,end_row=6+len(results.classes),end_column=curr_column)
            activity_column+=1

        ws.cell(3,start_column).value=days[day]
        ws.merge_cells(start_row=3,start_column=start_column,end_row=3,end_column=start_column+lesson_num+activity_num-1)

        ws.cell(4,start_column).value="上午"
        ws.merge_cells(start_row=4,start_column=start_column,end_row=4,end_column=start_column+morning_activity_num+morning_class_num-1)
        ws.cell(4,start_column+morning_activity_num+morning_class_num).value="下午"
        ws.merge_cells(start_row=4,start_column=start_column+morning_activity_num+morning_class_num,end_row=4,end_column=start_column+lesson_num+activity_num-1)

    for row in ws[ws.dimensions]:
        for cell in row:
            if cell.column>total_column:
                continue
            cell.alignment=center_align
            cell.border=border
            if cell.font!=title_font:
                cell.font=font

    wb.save(filename+"-班级总表（横）"+ext)
    logging.info("保存 班级总表（横） 成功")

def save_teachers_timetable(filename:str,ext:str):
    wb=Workbook()
    wb.remove(wb.active)
    for grade in grades:
        ws:openpyxl.worksheet.worksheet.Worksheet=wb.create_sheet(grade+"各教师课表")
        ws.column_dimensions["A"].width=cfg.text_size.value-2
        ws.column_dimensions["B"].width=cfg.text_size.value+2
        ws.column_dimensions["C"].width=ws.column_dimensions["D"].width=ws.column_dimensions["E"].width=ws.column_dimensions["F"].width=ws.column_dimensions["G"].width=cfg.text_size.value*1.5
        start_row=1
        for teacher_name in teachers[grade]:
            teacher=results.teachers[teacher_name]
            ws.cell(start_row,1).value=cfg.school_name.value
            ws.cell(start_row,1).font=title_font
            ws.cell(start_row,1).border=border
            ws.merge_cells(start_row=start_row,start_column=1,end_row=start_row,end_column=7)
            ws.row_dimensions[start_row].height=ws.row_dimensions[start_row+1].height=ws.row_dimensions[start_row+2].height=info_row_height
            ws.cell(start_row+1,1).value=teacher_name+"课表"
            ws.cell(start_row+1,1).font=title_font
            ws.cell(start_row+1,1).border=border
            ws.merge_cells(start_row=start_row+1,start_column=1,end_row=start_row+1,end_column=7)
            ws.append(["星期\n节次","","星期一","星期二","星期三","星期四","星期五"])
            for day in range(1,6):
                ws.cell(start_row+2,day+2).border=border
                for row in range(start_row+3,start_row+3+lesson_num+activity_num):
                    ws.row_dimensions[row].height=lesson_row_height

            for lesson in range(1,lesson_num+1):
                ws.cell(start_row+2+lesson,2).border=border
                ws.cell(start_row+2+lesson,2).value=f"{lesson}节\n%02d:%02d~%02d:%02d"%(cfg.lessons_time.value[str(lesson)][0][0],cfg.lessons_time.value[str(lesson)][0][1],cfg.lessons_time.value[str(lesson)][1][0],cfg.lessons_time.value[str(lesson)][1][1])

            timetable=teacher.timetable_dataframe
            for day in range(1,6):
                col=day+2
                for lesson in range(1,lesson_num+1):
                    row=start_row+2+lesson
                    time=Time(day,lesson)
                    subject_name=timetable.loc[time.to_str(False,True),time.to_str(True,False)]
                    ws.cell(row,col).value=subject_name
                    ws.cell(row,col).border=border

            activity_row=0
            for activity in list(activities.keys())+last_activities:
                if activity in activities:
                    lesson=activities[activity]
                else:
                    lesson=lesson_num+1
                curr_row=start_row+2+lesson+activity_row
                ws.insert_rows(curr_row)
                ws.cell(curr_row,2).value=activity+" %02d:%02d~%02d:%02d"%(cfg.activity_info.value[activity][0][0],cfg.activity_info.value[activity][0][1],cfg.activity_info.value[activity][1][0],cfg.activity_info.value[activity][1][1])
                ws.cell(curr_row,2).border=border
                ws.row_dimensions[curr_row].height=activity_row_height
                ws.merge_cells(start_row=curr_row,start_column=2,end_row=curr_row,end_column=7)
                activity_row+=1

            ws.merge_cells(start_row=start_row+2,start_column=1,end_row=start_row+2,end_column=2)
            ws.cell(start_row+3,1).value="上午"
            ws.cell(start_row+3,1).border=border
            ws.merge_cells(start_row=start_row+3,start_column=1,end_row=start_row+2+morning_class_num+morning_activity_num,end_column=1)
            afternoon_row=start_row+3+morning_class_num+morning_activity_num
            ws.cell(afternoon_row,1).value="下午"
            ws.cell(afternoon_row,1).border=border
            ws.merge_cells(start_row=afternoon_row,start_column=1,end_row=start_row+2+lesson_num+activity_num,end_column=1)

            start_row+=4+lesson_num+activity_num
        for row in ws[ws.dimensions]:
            for cell in row:
                cell.alignment=center_align
                if cell.font!=title_font:
                    cell.font=font

    wb.save(filename+"-教师课表（年级）"+ext)
    logging.info("保存 教师课表（年级） 成功")

def save_total_teachers_timetable(filename:str,ext:str):
    font=Font(cfg.text_font.value,cfg.text_size.value+4)
    total_column=2+(lesson_num+activity_num)*5
    wb=Workbook()
    ws=wb.active
    ws.title=os.path.basename(filename)+"_总表"
    ws["A1"].value=cfg.school_name.value
    ws["A1"].font=title_font
    ws.row_dimensions[1].height=ws.row_dimensions[2].height=info_row_height
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=total_column)

    ws["A2"].value="总课表（教师）"
    ws["A2"].font=title_font
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=total_column)

    ws["A3"].value="教师\n节次\n时间"
    ws.merge_cells("A3:A6")

    for day in range(1,6):
        start_row=7
        start_column=(day-1)*(lesson_num+activity_num)+2
        teachers=list(results.teachers.values())
        for i in range(len(teachers)):
            teacher=teachers[i]
            ws.cell(start_row+i,1).value=teacher.name
            timetable=teacher.timetable_dataframe

            for lesson in range(1,lesson_num+activity_num+1):
                if lesson<=lesson_num:
                    ws.cell(5,start_column+lesson-1).value=f"{lesson}节"
                    ws.cell(6,start_column+lesson-1).value="%02d:%02d~%02d:%02d"%(cfg.lessons_time.value[str(lesson)][0][0],cfg.lessons_time.value[str(lesson)][0][1],cfg.lessons_time.value[str(lesson)][1][0],cfg.lessons_time.value[str(lesson)][1][1])
                    time=Time(day,lesson)
                    subject_name=timetable.loc[time.to_str(False,True),time.to_str(True,False)]
                    ws.cell(start_row+i,start_column+lesson-1).value=subject_name
                ws.column_dimensions[column_letter(start_column+lesson-1)].width=cfg.text_size.value+5


        activity_column=0
        for activity in list(activities.keys())+last_activities:
            if activity in activities:
                lesson=activities[activity]
            else:
                lesson=lesson_num+1
            curr_column=start_column+lesson+activity_column-1
            ws.insert_cols(curr_column)
            ws.cell(5,curr_column).value="%02d:%02d~%02d:%02d"%(cfg.activity_info.value[activity][0][0],cfg.activity_info.value[activity][0][1],cfg.activity_info.value[activity][1][0],cfg.activity_info.value[activity][1][1])
            ws.merge_cells(start_row=5,start_column=curr_column,end_row=6,end_column=curr_column)

            ws.cell(7,curr_column).value=activity
            ws.merge_cells(start_row=7,start_column=curr_column,end_row=6+len(results.teachers),end_column=curr_column)
            activity_column+=1

        ws.cell(3,start_column).value=days[day]
        ws.merge_cells(start_row=3,start_column=start_column,end_row=3,end_column=start_column+lesson_num+activity_num-1)

        ws.cell(4,start_column).value="上午"
        ws.merge_cells(start_row=4,start_column=start_column,end_row=4,end_column=start_column+morning_activity_num+morning_class_num-1)
        ws.cell(4,start_column+morning_activity_num+morning_class_num).value="下午"
        ws.merge_cells(start_row=4,start_column=start_column+morning_activity_num+morning_class_num,end_row=4,end_column=start_column+lesson_num+activity_num-1)

    for row in ws[ws.dimensions]:
        for cell in row:
            if cell.column>total_column:
                continue
            cell.alignment=center_align
            cell.border=border
            if cell.font!=title_font:
                cell.font=font

    wb.save(filename+"-教师总表（横）"+ext)
    logging.info("保存 教师总表（横） 成功")

class SaveThread(QThread):
    success=Signal()
    error=Signal(str)

    def __init__(self,filename: str,ext: str,parent=None):
        super().__init__(parent)
        self.filename=filename
        self.ext=ext

    def run(self):
        try:
            save_grades_timetable(self.filename,self.ext)
            save_total_grades_timetable(self.filename,self.ext)
            save_teachers_timetable(self.filename,self.ext)
            save_total_teachers_timetable(self.filename,self.ext)
            self.success.emit()
        except Exception as estr:
            e=traceback.format_exc()
            logging.critical(f"保存课程表出错：\n{e}")
            self.error.emit(str(estr))